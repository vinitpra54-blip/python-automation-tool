import os 
from os.path import abspath
import sys

from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class ExcelLibary:

    def __init__(self):
        self.lable = {
            "TestResult":[
                "TestCase Name"
                ,"Obective"
                ,"Test Step"
                ,"Expected Result"
                ,"Result"
                ,"Date Execute"
            ],
            "CoverPage":[
                "Test Date Time"
                ,"Test By"
                ,"Env. Test"
            ]
        }
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

     
    def _apply_styles(self, ws_testResult, all_rows_data):     

        for row_index in range(len(all_rows_data)):
            row_index = row_index +1

            current_row = ws_testResult[row_index]
            is_header_row = (row_index == 1)

            for cell in current_row:
                cell.border = self.thin_border
                
                if is_header_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                else:
                    cell.alignment = self.data_alignment

    def _set_column_widths(self, ws_testResult):
        ws_testResult.column_dimensions['A'].width = 20
        ws_testResult.column_dimensions['B'].width = 40
        ws_testResult.column_dimensions['C'].width = 50
        ws_testResult.column_dimensions['D'].width = 60
        ws_testResult.column_dimensions['E'].width = 10
        ws_testResult.column_dimensions['F'].width = 20

    def sheet_CoverPage(self,all_rows_data):
        ws_testResult = self.wb['Sheet']
        ws_testResult.title = "CoverPage"

        all_row_datas = list(zip(*all_rows_data))

        for index_row,row_data in enumerate(all_row_datas):
            index_row = index_row + 1 

            for index_col ,cell_value in enumerate(row_data):
                index_col = index_col + 1
                ws_testResult.cell(row=index_row , column=index_col , value= cell_value)

            for cell in ws_testResult.iter_rows(min_row=index_row,max_row=index_row,min_col=1,max_col=1 ):
                for cells in cell :
                    cells.border = self.thin_border
                    cells.font = self.header_font
                    cells.fill = self.header_fill

            for cell in ws_testResult.iter_rows(min_row=index_row,max_row=index_row,min_col=1,max_col=2 ):
                for cells in cell :
                    cells.border = self.thin_border

        ws_testResult.column_dimensions['A'].width = 25
        ws_testResult.column_dimensions['B'].width = 25
        return True

    def sheet_TestResultOfLoginWeb(self, all_rows_data): 
        ws_testResult = self.wb.create_sheet("TestResultOfLoginWeb") 

        for row_data in all_rows_data:
            ws_testResult.append(row_data)

        self._apply_styles(ws_testResult, all_rows_data)
        self._set_column_widths(ws_testResult)

    def sheet_TestResultOfApi(self, all_rows_data): 
        ws_testResult = self.wb.create_sheet("TestResultOfApi") 
        
        for row_data in all_rows_data:
            ws_testResult.append(row_data)

        self._apply_styles(ws_testResult, all_rows_data)
        self._set_column_widths(ws_testResult)

    def saveTestResultOfLoginWeb(self, listResult):
       
        all_rows_data = []
        header_row = self.lable["TestResult"]
        all_rows_data.append(header_row)
        
        for test_case_block in listResult:
            if test_case_block and test_case_block[0]:
                data_row = test_case_block[0]
                all_rows_data.append(data_row)

        self.sheet_TestResultOfLoginWeb(all_rows_data) 
        return True
    
    def saveTestResultOfApi(self, listResult):
       
        all_rows_data = []
        header_row = self.lable["TestResult"]
        all_rows_data.append(header_row)
        
        for test_case_block in listResult:
            if test_case_block and test_case_block[0]:
                data_row = test_case_block[0]
                all_rows_data.append(data_row)

        self.sheet_TestResultOfApi(all_rows_data) 

        return True

    def saveTestResultOfApi(self, listResult):
       
        all_rows_data = []
        header_row = self.lable["TestResult"]
        all_rows_data.append(header_row)
        
        for test_case_block in listResult:
            if test_case_block and test_case_block[0]:
                data_row = test_case_block[0]
                all_rows_data.append(data_row)

        self.sheet_TestResultOfApi(all_rows_data) 

        return True
    
    def saveCoverPage(self, listResult):
        all_rows_data = []
        header_row = self.lable["CoverPage"]
        all_rows_data.append(header_row)
   
        for test_case_block in listResult:
            data_row = test_case_block
            all_rows_data.append(data_row)
        self.sheet_CoverPage(all_rows_data) 

        return True
    
    def writeExcel(self,filePath,coverPage,listResult,listResult_1):

        self.wb = Workbook()
        try:
            self.saveCoverPage(coverPage)
            self.saveTestResultOfLoginWeb(listResult)
            self.saveTestResultOfApi(listResult_1)
            self.wb.save(str(filePath))
            self.wb.close()
            print(f'Excel Result {filePath} has been written successfully')
        except Exception as e :
            return e   
    
        return True   
        
def main():
    obj = ExcelLibary()

    coverPage = [['20251208_154937', 'Vinit Prasobphon', 'Dev1']]

    listResult_1=[
            [
                [
                    "Login success",
                    "verify that users can login successfully when input a correct username and password.",
                    "1. Open browser and go to http://theinternet.herokuapp.com/login.  \n2. Input username tomsmith and password  SuperSecretPassword!.  \n3. Click the Logout button.\n",
                    "1. Login page is shown. \n2. Login success and message You logged into a secure area! is shown.  \n3. Go back to the Login page and the message  You logged out of the secure area! is shown.\n",
                    "Pass",
                    "20251129_171428"
                ]
            ],
            [
                [
                    "Login failed - Password incorrect",
                    "verify that users can login unsuccessfully when they input a correct username but wrong password.",
                    "1. Open browser and go to http://theinternet.herokuapp.com/login.  \n2. Input username tomsmith and password Password!.\n",
                    "1. Login page is shown. \n2. Login failed and the message Your password is invalid! is shown.\n",
                    "Pass",
                    "20251129_171438"
                ]
            ]
    ]

    listResult=[
            [
                [
                    "Login success",
                    "verify that users can login successfully when input a correct username and password.",
                    "1. Open browser and go to http://theinternet.herokuapp.com/login.  \n2. Input username tomsmith and password  SuperSecretPassword!.  \n3. Click the Logout button.\n",
                    "1. Login page is shown. \n2. Login success and message You logged into a secure area! is shown.  \n3. Go back to the Login page and the message  You logged out of the secure area! is shown.\n",
                    "Pass",
                    "20251129_171428"
                ]
            ],
            [
                [
                    "Login failed - Password incorrect",
                    "verify that users can login unsuccessfully when they input a correct username but wrong password.",
                    "1. Open browser and go to http://theinternet.herokuapp.com/login.  \n2. Input username tomsmith and password Password!.\n",
                    "1. Login page is shown. \n2. Login failed and the message Your password is invalid! is shown.\n",
                    "Pass",
                    "20251129_171438"
                ]
            ],
            [
                [
                    "Login failed - Username not found incorrect",
                    "verify that users can login unsuccessfully when they input a username that did not exist.",
                    "1. Open browser and go to http://theinternet.herokuapp.com/login.  \n2. Input username tomholland and password Password!.\n",
                    "1. Login page is shown.  \n2. Login failed and the message Your username is invalid! is shown.\n",
                    "Pass",
                    "20251129_171447"
                ]
            ]
        ]
    # obj.saveTestResult(listResult,'Test_Automation_Results.xlsx')
    obj.writeExcel('Test_Automation_Results.xlsx',coverPage,listResult,listResult_1)


if __name__ == '__main__':
    main()
