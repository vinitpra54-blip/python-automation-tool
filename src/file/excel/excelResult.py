import os 
from os.path import abspath
import sys

from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class ExcelLibary:


    def __init__(self):
        self.lable = {
            "TestResult":[
                "TestCase Name"
                ,"Obective"
                ,"Test Step"
                ,"Expected Result"
                ,"Result"
                ,"Date Execute"
            ]
        }
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

     
    def _apply_styles(self, ws_testResult, all_rows_data):     
        for row_index, row_data in enumerate(all_rows_data, 1):
            current_row = ws_testResult[row_index]
            is_header_row = (row_index == 1)

            for cell in current_row:
                cell.border = self.thin_border
                
                if is_header_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                else:
                    cell.alignment = self.data_alignment

    def _set_column_widths(self, ws_testResult, all_rows_data):
        if all_rows_data:
            # Widths: [TestCase Name, Obective, Test Step, Expected Result, Result, Date Execute]
            widths = [20, 40, 50, 60, 10, 20] 
            
            num_cols = len(all_rows_data[0])
            for col_index in range(1, num_cols + 1):
                col_letter = get_column_letter(col_index)
                if col_index <= len(widths):
                    ws_testResult.column_dimensions[col_letter].width = widths[col_index - 1]

    def sheet_TestResult(self, all_rows_data): 
        ws_testResult = self.wb.active 
        ws_testResult.title = "TestResult"
        
        for row_data in all_rows_data:
            ws_testResult.append(row_data)

        self._apply_styles(ws_testResult, all_rows_data)
        self._set_column_widths(ws_testResult, all_rows_data)

    def saveTestResult(self, listResult, filePath):
        self.wb = Workbook()
        try:
    
            all_rows_data = []
            header_row = self.lable["TestResult"]
            all_rows_data.append(header_row)
            
            for test_case_block in listResult:
                if test_case_block and test_case_block[0]:
                    data_row = test_case_block[0]
                    all_rows_data.append(data_row)
            
            self.sheet_TestResult(all_rows_data) 

            
            self.wb.save(str(filePath))
            self.wb.close() 
            
            print(f'Excel Result {filePath} has been written successfully')
            return f'Excel Result {filePath} has been written successfully'
            
        except Exception as e :
            return e
        

def main():
    obj = ExcelLibary()

    listResult =[
        [
            [
                "Login success",
                "verify that users can login successfully when input a correct username and password.",
                "1. Open browser and go to http://theinternet.herokuapp.com/login. 2. Input username tomsmith and password  SuperSecretPassword!. /n 3. Click the Logout button. ",
                "1. Login page is shown. /n 2. Login success and message You logged into a secure area! is shown. /n 3. Go back to the Login page and the message  You logged out of the secure area! is shown.",
                "Pass",
                "20251128_122114"
            ]
        ],
        [
            [
                "Login failed - Password",
                "verify that users can login unsuccessfully when they input a correct username but wrong password.",
                "1. Open browser and go to http://theinternet.herokuapp.com/login.  2. Input username tomsmith and password Password!.",
                "1. Login page is shown. 2. Login failed and the message Your password is invalid! is shown.",
                "Pass",
                "20251128_122130"
            ]
        ],
        [
            [
                "Login failed - Username not found incorrect",
                "verify that users can login unsuccessfully when they input a username that did not exist.",
                "1. Open browser and go to http://theinternet.herokuapp.com/login.  2. Input username tomholland and password Password!.",
                "1. Login page is shown.  2. Login failed and the message Your username is invalid! is shown. ",
                "Pass",
                "20251128_122145"
            ]
        ]
    ]
    obj.saveTestResult(listResult,'Test_Automation_Results.xlsx')
    # obj.sheet_TestResult(listResult)


if __name__ == '__main__':
    main()
